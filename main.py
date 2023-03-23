# Program do porównywania db dumpa oraz bazy danych dostarczonej przez stocznię.

import openpyxl
# from openpyxl.comments import Comment
# from openpyxl.styles import PatternFill
# import pandas as pd

import tkinter
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox

from Functions import *


class CreateToolTip(object):
    def __init__(self, widget, text='widget info'):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.close)

    def enter(self, event=None):
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        # creates a toplevel window
        self.tw = tkinter.Toplevel(self.widget)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = tkinter.Label(self.tw, text=self.text, justify='left', relief='solid', borderwidth=1,
                              font=("times", "15", "normal"))
        label.pack(ipadx=1)

    def close(self, event=None):
        if self.tw:
            self.tw.destroy()


# Graphic animation function
root = Tk(className=' CdA DataBase Comparator')
root.geometry('790x185')
root.resizable(False, False)
# Variables
# File 1 Variables

file_path1 = ''
file_path_txt1 = StringVar()
file_path_txt1.set('Select import file to correct')
wb1 = openpyxl.Workbook()
active_sheet1 = wb1.active
max_rows1 = 0
max_col1 = 0

file1_col_tag = 0
file1_col_template = 0
file1_col_event = 0
file1_col_instrument_code = 0
file1_col_seq = 0
file1_col_msggroup = 0
file1_col_algroup = 0
file1_col_prio = 0
file1_col_limit1 = 0
file1_col_limit2 = 0
file1_col_limit3 = 0
file1_col_limit4 = 0
file1_col_HHprio = 0
file1_col_Hprio = 0
file1_col_Lprio = 0
file1_col_LLprio = 0
file1_col_Ext1prio = 0
file1_col_Ext2prio = 0
file1_col_Ext3prio = 0
file1_col_Ext4prio = 0
file1_col_FAprio = 0
file1_col_HHca = 0                  # HH alarm and control
file1_col_Hca = 0                   # H alarm and control
file1_col_Lca = 0                   # L alarm and control
file1_col_LLca = 0                  # LL alarm and control
file1_col_devicetag1 = 0
file1_col_devicetag2 = 0
file1_col_devicetag3 = 0
file1_col_devicetag4 = 0
file1_col_devicetag5 = 0
file1_col_devicetag6 = 0
file1_col_devicetag7 = 0
file1_col_devicetag8 = 0
file1_col_devicetag9 = 0
file1_col_devicetag10 = 0

# File 2 Variables
file_path2 = ''
file_path_txt2 = StringVar()
file_path_txt2.set('System - Alarm mapping table')
wb2 = openpyxl.Workbook()
active_sheet2 = wb2.active
max_rows2 = 0
max_col2 = 0


def analyze_file():
    global active_sheet1
    global max_rows1
    global max_col1
    global active_sheet2
    global max_rows2
    global max_col2

    wb1.active = wb1[sheet_choose1.get()]
    active_sheet1 = wb1.active
    max_rows1 = wb1.active.max_row
    max_col1 = wb1.active.max_column
    wb2.active = wb2[sheet_choose2.get()]
    active_sheet2 = wb2.active
    max_rows2 = wb2.active.max_row
    max_col2 = wb2.active.max_column
    fill_col_numbers()
    print(file_path1)
    print(file_path2)

    if checkbox2_var.get() == 1:
        am100_clean_devicetags(active_sheet1, file1_col_template, file1_col_devicetag2, file1_col_devicetag3,
                               file1_col_devicetag4, file1_col_devicetag5, file1_col_devicetag6, file1_col_devicetag7,
                               file1_col_devicetag8, file1_col_devicetag9, file1_col_devicetag10, max_rows1)
        am100_update_devicetags(active_sheet1, file1_col_tag, file1_col_seq, file1_col_prio, file1_col_template,
                                file1_col_instrument_code,file1_col_devicetag1, file1_col_devicetag2,
                                file1_col_devicetag3, file1_col_devicetag4, file1_col_devicetag5, file1_col_devicetag6,
                                file1_col_devicetag7, file1_col_devicetag8, file1_col_devicetag9, file1_col_devicetag10,
                                file1_col_HHprio, file1_col_Hprio, file1_col_Lprio, file1_col_LLprio, file1_col_FAprio,
                                file1_col_Ext1prio, file1_col_Ext2prio, file1_col_Ext3prio, file1_col_Ext4prio, max_rows1)
        wb1.save(file_path1)
    if checkbox1_var.get() == 1:
        am100_active = 0
        if checkbox2_var.get() == 1:
            am100_active = 1
        update_alm_and_msg(active_sheet1, active_sheet2, am100_active, file1_col_tag, file1_col_seq, file1_col_template,
                           file1_col_event, file1_col_algroup, file1_col_msggroup, file1_col_prio, file1_col_HHprio,
                           file1_col_Hprio, file1_col_Lprio, file1_col_LLprio, file1_col_Ext1prio, file1_col_Ext2prio,
                           file1_col_Ext3prio,file1_col_Ext3prio,file1_col_FAprio, max_rows1, max_rows2)
    if checkbox3_var.get() == 1:
        merge_alarms(active_sheet1, file1_col_template, file1_col_seq, file1_col_HHca, file1_col_Hca, file1_col_Lca,
                     file1_col_LLca, file1_col_limit1, file1_col_limit2, file1_col_limit3, file1_col_limit4, max_rows1)

    if checkbox4_var.get() == 1:
        print('Done4')

    wb1.save(file_path1)
    wb2.save(file_path2)

    messagebox.showinfo(title='Message', message='Finished')


def fill_col_numbers():
    # DNA DUMP
    global file1_col_tag
    global file1_col_template
    global file1_col_event
    global file1_col_instrument_code
    global file1_col_seq
    global file1_col_msggroup
    global file1_col_algroup
    global file1_col_prio
    global file1_col_limit1
    global file1_col_limit2
    global file1_col_limit3
    global file1_col_limit4
    global file1_col_HHprio
    global file1_col_Hprio
    global file1_col_Lprio
    global file1_col_LLprio
    global file1_col_Ext1prio
    global file1_col_Ext2prio
    global file1_col_Ext3prio
    global file1_col_Ext4prio
    global file1_col_FAprio
    global file1_col_HHca
    global file1_col_Hca
    global file1_col_Lca
    global file1_col_LLca
    global file1_col_devicetag1
    global file1_col_devicetag2
    global file1_col_devicetag3
    global file1_col_devicetag4
    global file1_col_devicetag5
    global file1_col_devicetag6
    global file1_col_devicetag7
    global file1_col_devicetag8
    global file1_col_devicetag9
    global file1_col_devicetag10
    global max_col1

    file1_col_tag = get_col_no(active_sheet1, '$(LOOP)', max_col1)
    file1_col_template = get_col_no(active_sheet1, '$(TEMPLATE)', max_col1)
    file1_col_event = get_col_no(active_sheet1, '$(EVENT)', max_col1)
    file1_col_instrument_code = get_col_no(active_sheet1, '$(INSTRUMENT_CODE)', max_col1)
    file1_col_seq = get_col_no(active_sheet1, '$(LOOP_ORDER)', max_col1)
    file1_col_msggroup = get_col_no(active_sheet1, '$(MESGROUP)', max_col1)
    file1_col_algroup = get_col_no(active_sheet1, '$(ALGROUP)', max_col1)
    file1_col_prio = get_col_no(active_sheet1, '$(ALPRI1)', max_col1)
    file1_col_limit1 = get_col_no(active_sheet1, '$(LIMIT1)', max_col1)
    file1_col_limit2 = get_col_no(active_sheet1, '$(LIMIT2)', max_col1)
    file1_col_limit3 = get_col_no(active_sheet1, '$(LIMIT3)', max_col1)
    file1_col_limit4 = get_col_no(active_sheet1, '$(LIMIT4)', max_col1)
    file1_col_HHprio = get_col_no(active_sheet1, '$(ALPRIHH)', max_col1)
    file1_col_Hprio = get_col_no(active_sheet1, '$(ALPRIH)', max_col1)
    file1_col_Lprio = get_col_no(active_sheet1, '$(ALPRIL)', max_col1)
    file1_col_LLprio = get_col_no(active_sheet1, '$(ALPRILL)', max_col1)
    file1_col_Ext1prio = get_col_no(active_sheet1, '$(ALPRIEXT1)', max_col1)
    file1_col_Ext2prio = get_col_no(active_sheet1, '$(ALPRIEXT2)', max_col1)
    file1_col_Ext3prio = get_col_no(active_sheet1, '$(ALPRIEXT3)', max_col1)
    file1_col_Ext4prio = get_col_no(active_sheet1, '$(ALPRIEXT4)', max_col1)
    file1_col_FAprio = get_col_no(active_sheet1, '$(ALPRIFA)', max_col1)            #sprawdzic czy tak sie nazywa
    file1_col_HHca = get_col_no(active_sheet1, '$(HH_C&A)', max_col1)
    file1_col_Hca = get_col_no(active_sheet1, '$(H_C&A)', max_col1)
    file1_col_Lca = get_col_no(active_sheet1, '$(L_C&A)', max_col1)
    file1_col_LLca = get_col_no(active_sheet1, '$(LL_C&A)', max_col1)
    file1_col_devicetag1 = get_col_no(active_sheet1, '$(DEVICETAG1)', max_col1)
    file1_col_devicetag2 = get_col_no(active_sheet1, '$(DEVICETAG2)', max_col1)
    file1_col_devicetag3 = get_col_no(active_sheet1, '$(DEVICETAG3)', max_col1)
    file1_col_devicetag4 = get_col_no(active_sheet1, '$(DEVICETAG4)', max_col1)
    file1_col_devicetag5 = get_col_no(active_sheet1, '$(DEVICETAG5)', max_col1)
    file1_col_devicetag6 = get_col_no(active_sheet1, '$(DEVICETAG6)', max_col1)
    file1_col_devicetag7 = get_col_no(active_sheet1, '$(DEVICETAG7)', max_col1)
    file1_col_devicetag8 = get_col_no(active_sheet1, '$(DEVICETAG8)', max_col1)
    file1_col_devicetag9 = get_col_no(active_sheet1, '$(DEVICETAG9)', max_col1)
    file1_col_devicetag10 = get_col_no(active_sheet1, '$(DEVICETAG10)', max_col1)

    if checkbox2_var.get() == 1 and file1_col_Ext1prio == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(ALPRIEXT1)', 1)
        file1_col_Ext1prio = max_col1 + 2
        set_cell_value(active_sheet1, 1, max_col1 + 2, '$(ALPRIEXT2)', 1)
        file1_col_Ext2prio = max_col1 + 2
        set_cell_value(active_sheet1, 1, max_col1 + 3, '$(ALPRIEXT3)', 1)
        file1_col_Ext3prio = max_col1 + 3
        set_cell_value(active_sheet1, 1, max_col1 + 4, '$(ALPRIEXT4)', 1)
        file1_col_Ext4prio = max_col1 + 4
        set_cell_value(active_sheet1, 1, max_col1 + 5, '$(ALPRIFA)', 1)
        file1_col_FAprio = max_col1 + 5
        max_col1 = max_col1 + 5

    if checkbox2_var.get() == 1 and (file1_col_devicetag1 == 0 or file1_col_devicetag2 == 0
                                     or file1_col_devicetag3 == 0 or file1_col_devicetag4 == 0
                                     or file1_col_devicetag5 == 0 or file1_col_devicetag6 == 0
                                     or file1_col_devicetag7 == 0 or file1_col_devicetag8 == 0
                                     or file1_col_devicetag9 == 0 or file1_col_devicetag10 == 0):
        if file1_col_devicetag1 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG1)', 1)
            max_col1 += 1
            file1_col_devicetag1 = max_col1
        if file1_col_devicetag2 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG2)', 1)
            max_col1 += 1
            file1_col_devicetag2 = max_col1
        if file1_col_devicetag3 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG3)', 1)
            max_col1 += 1
            file1_col_devicetag3 = max_col1
        if file1_col_devicetag4 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG4)', 1)
            max_col1 += 1
            file1_col_devicetag4 = max_col1
        if file1_col_devicetag5 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG5)', 1)
            max_col1 += 1
            file1_col_devicetag5 = max_col1
        if file1_col_devicetag6 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG6)', 1)
            max_col1 += 1
            file1_col_devicetag6 = max_col1
        if file1_col_devicetag7 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG7)', 1)
            max_col1 += 1
            file1_col_devicetag7 = max_col1
        if file1_col_devicetag8 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG8)', 1)
            max_col1 += 1
            file1_col_devicetag8 = max_col1
        if file1_col_devicetag9 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG9)', 1)
            max_col1 += 1
            file1_col_devicetag9 = max_col1
        if file1_col_devicetag10 == 0:
            set_cell_value(active_sheet1, 1, max_col1 + 1, '$(DEVICETAG10)', 1)
            max_col1 += 1
            file1_col_devicetag10 = max_col1
    if file1_col_HHprio == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(ALPRIHH)', 1)
        max_col1 += 1
        file1_col_HHprio = max_col1
    if file1_col_Hprio == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(ALPRIH)', 1)
        max_col1 += 1
        file1_col_Hprio = max_col1
    if file1_col_Lprio == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(ALPRIL)', 1)
        max_col1 += 1
        file1_col_Lprio = max_col1
    if file1_col_LLprio == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(ALPRILL)', 1)
        max_col1 += 1
        file1_col_LLprio = max_col1
    if file1_col_event == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(EVENT)', 1)
        max_col1 += 1
        file1_col_event = max_col1
    if file1_col_limit1 == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(LIMIT1)', 1)
        max_col1 += 1
        file1_col_limit1 = max_col1
    if file1_col_limit2 == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(LIMIT2)', 1)
        max_col1 += 1
        file1_col_limit2 = max_col1
    if file1_col_limit3 == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(LIMIT3)', 1)
        max_col1 += 1
        file1_col_limit3 = max_col1
    if file1_col_limit4 == 0:
        set_cell_value(active_sheet1, 1, max_col1 + 1, '$(LIMIT4)', 1)
        max_col1 += 1
        file1_col_limit4 = max_col1

def choose_file1():
    global file_path1
    global wb1
    root.filename = filedialog.askopenfilename(title='Choose file to open',
                                               filetypes=(('xlsx', '*.xlsx'), ('xls', '*.xls')))
    if len(root.filename) > 0:
        file_path1 = root.filename
        file_path_txt1.set('File: ' + file_path1)
    if len(file_path1) > 0:
        wb1 = openpyxl.load_workbook(file_path1)
        sheet_names1 = [wb1.sheetnames]
        sheet_choose1['values'] = tuple(sheet_names1[0])
        sheet_choose1.current(0)
        if len(file_path2) > 2:
            button_analyze['state'] = tkinter.NORMAL


def choose_file2():
    global file_path2
    global wb2
    root.filename = filedialog.askopenfilename(title='Choose file to open',
                                               filetypes=(('xlsx', '*.xlsx'), ('xls', '*.xls')))
    if len(root.filename) > 0:
        file_path2 = root.filename
        file_path_txt2.set('File: ' + file_path2)
    if len(file_path2) > 0:
        wb2 = openpyxl.load_workbook(file_path2)
        sheet_names2 = [wb2.sheetnames]
        sheet_choose2['values'] = tuple(sheet_names2[0])
        sheet_choose2.current(0)
        if len(file_path1) > 0:
            button_analyze['state'] = tkinter.NORMAL
# ------- Graphic user interface --------
# ---------------------------------------


file_label1 = Label(root, textvariable=file_path_txt1, width=50, anchor='w', relief='groove')
file_label1.place(x=10, y=20)
sheet_choose_select1 = tkinter.StringVar()
sheet_choose1 = ttk.Combobox(root, textvariable=sheet_choose_select1, width=10, height=1)
sheet_choose1.place(x=470, y=17)
button_select1 = Button(root, text='Select', command=choose_file1, height=1, width=5)
button_select1.place(x=590, y=16)

button_select2 = Button(root, text='Select', command=choose_file2, height=1, width=5)
button_select2.place(x=590, y=46)
file_label2 = Label(root, textvariable=file_path_txt2, width=50, anchor='w', relief='groove')

file_label2.place(x=10, y=49)
sheet_choose_select2 = tkinter.StringVar()
sheet_choose2 = ttk.Combobox(root, textvariable=sheet_choose_select2, width=10, height=1)
sheet_choose2.place(x=470, y=47)

button_analyze = Button(root, text='Analyze', command=analyze_file, height=3, width=8, state=tkinter.DISABLED)
button_analyze.place(x=675, y=17)

labelframe1 = ttk.Labelframe(root, width=770, height=95, labelanchor=NW, text='Check options')
labelframe1.place(x=10, y=80)

checkbox1_var = IntVar(root, 1)
checkbox1 = Checkbutton(root, text='Update Al & Msg group', variable=checkbox1_var, onvalue=1,
                        offvalue=0, height=1)
checkbox1.place(x=20, y=105)
checkbox1_tt = CreateToolTip(checkbox1, 'Update alarm and message groups according to given table')

checkbox2_var = IntVar(root, 0)
checkbox2 = Checkbutton(root, text='Update Am100 alarm & connections', variable=checkbox2_var, onvalue=1,
                        offvalue=0, height=1)
checkbox2.place(x=20, y=127)
checkbox2_tt = CreateToolTip(checkbox2, 'Update Am100 Alarms and EXT1..4 connections with alarm prio')

checkbox3_var = IntVar(root, 0)
checkbox3 = Checkbutton(root, text='Merge Alarm and Control limits', variable=checkbox3_var, onvalue=1,
                        offvalue=0, height=1)
checkbox3.place(x=20, y=149)
checkbox3_tt = CreateToolTip(checkbox3, 'Merging alarm and control limits for Am templates if need')

checkbox4_var = IntVar(root, 0)
checkbox4 = Checkbutton(root, text='Spare', variable=checkbox4_var, onvalue=1,
                        offvalue=0, height=1)
checkbox4.place(x=260, y=105)
checkbox4_tt = CreateToolTip(checkbox4, 'New function')
root.mainloop()
