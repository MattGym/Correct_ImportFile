from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


def get_col_no(sheet, desc, max_col):
    """
    According to given active_sheet name and searched column description (str), function
    returns column number where that string is.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    desc : str
        searched string
    max_col : int
        number of not empty columns
    """
    for s in range(max_col):
        if str(desc) == str(sheet.cell(row=1, column=s+1).value):
            return s+1
    return 0


def set_cell_color(sheet, row, col, color_rgn='n'):
    """
    Change cell background color on 'r'-red, 'g'-green, 'n'-none.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    color_rgn : str, optional
        (OPTIONAL) color_rgn = 'r'-red; 'g'-green; 'n' or empty - NONE
    """
    if color_rgn == 'r':
        fg_color = 'E74C3C'
        pattern = PatternFill(patternType='solid', fgColor=fg_color)
        sheet.cell(row=row, column=col).fill = pattern
    if color_rgn == 'g':
        fg_color = '2ECC71'
        pattern = PatternFill(patternType='solid', fgColor=fg_color)
        sheet.cell(row=row, column=col).fill = pattern
    if color_rgn == 'n':
        pattern = PatternFill(patternType=None)
        sheet.cell(row=row, column=col).fill = pattern


def get_cell_value(sheet, row, col):
    """
    Returns value from cell in active_sheet at specific row & column position.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    """
    return sheet.cell(row=row, column=col).value


def set_cell_value(sheet: object, row: object, col: object, val: object, typ: object = 0) -> object:
    """
    Set value at cell in active_sheet at specific row & column position.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    val : any
        value
    typ : int, optional
        (OPTIONAL) type of data (0 or empty - float/int; 1 - force string; 2 - None)
    """
    if typ == 0:
        sheet.cell(row=row, column=col).value = float(val)
    elif typ == 1:
        sheet.cell(row=row, column=col).value = str(val)
    elif typ == 2:
        sheet.cell(row=row, column=col).value = None
    elif typ == 3:
        sheet.cell(row=row, column=col).value = int(val)


def set_cell_comment(sheet, row, col, commentary, add=False, delete=False):
    """
    Function add 'commentary' to specified cell given as (active_sheet, row & column position)
    If optional parameter add=True then function add commentary to existing one.
    In another way removes old commentary and add a new one.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    commentary : str
        commentary that will be added to specified cell
    add : bool, optional
        (OPTIONAL) add=False or NONE - function swap commentary into new on , add=True - add second commentary
    delete : bool, optional
        (OPTIONAL) delete=False or NONE - function do nothing with existing commentary, delete=True - remove commentary
    """
    if str(sheet.cell(row=row, column=col).comment) == 'None' or add is False:
        comment = Comment(commentary, 'CdA analyzer')
        comment.width = 400
        comment.height = 150
        sheet.cell(row=row, column=col).comment = comment
    else:
        tmp_txt1 = str(str(sheet.cell(row=row, column=col).comment).replace('Comment: ', '')).\
            replace('by CdA analyzer', '')
        comment = Comment(tmp_txt1 + ' ::\n' + commentary, 'CdA analyzer')
        comment.width = 400
        comment.height = 150
        sheet.cell(row=row, column=col).comment = comment
    if delete is True:
        sheet.cell(row=row, column=col).comment = None


def update_alm_and_msg(sheet1, sheet2, am100_av, file1_col_tag, file1_col_seq,  file1_col_template, file1_col_event,
                       file1_col_alm, file1_col_msg, file1_col_prio, file1_col_hhprio, file1_col_hprio, file1_col_lprio,
                       file1_col_llprio, file1_col_ext1prio, file1_col_ext2prio, file1_col_ext3prio, file1_col_ext4prio,
                       file1_col_faprio, max_row1, max_row2):
    """
    Function 'update_alp_and_msg' update alarm and message group of core loop, according to given relationship table.
    ----------
    """
    for i in range(2, max_row1 + 1):
        if get_cell_value(sheet1, i, file1_col_seq) == 0:
            prefix = str(get_cell_value(sheet1, i, file1_col_tag))[0:3]
            for j in range(2, max_row2+1):
                if str(get_cell_value(sheet1, i, file1_col_alm)) == str(get_cell_value(sheet2, j, 1)) \
                        and (prefix == str(get_cell_value(sheet2, j, 2)) or get_cell_value(sheet2, j, 2) is None):
                    if str(get_cell_value(sheet1, i, file1_col_template)) == 'Am100' and am100_av is True:
                        print('Am100')
                        if get_cell_value(sheet1, i, file1_col_hhprio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_hprio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_lprio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_llprio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_ext1prio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_ext2prio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_ext3prio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_ext4prio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_faprio) == 900:
                            set_cell_value(sheet1, i, file1_col_alm, get_cell_value(sheet2, j, 3))
                            set_cell_value(sheet1, i, file1_col_msg, get_cell_value(sheet2, j, 3))
                            continue
                        set_cell_value(sheet1, i, file1_col_alm, get_cell_value(sheet2, j, 4))
                        set_cell_value(sheet1, i, file1_col_msg, get_cell_value(sheet2, j, 4))
                        continue
                    if str(get_cell_value(sheet1, i, file1_col_template))[0:4] == 'Am10' \
                            and str(get_cell_value(sheet1, i, file1_col_template)) != 'Am100':
                        print(str(get_cell_value(sheet1, i, file1_col_template))[0:4], '  ',               #do wyjebania
                              str(get_cell_value(sheet1, i, file1_col_template)))
                        if get_cell_value(sheet1, i, file1_col_hhprio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_hprio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_lprio) == 900 \
                                or get_cell_value(sheet1, i, file1_col_llprio) == 900:
                            set_cell_value(sheet1, i, file1_col_alm, get_cell_value(sheet2, j, 3))
                            set_cell_value(sheet1, i, file1_col_msg, get_cell_value(sheet2, j, 3))
                            continue
                        set_cell_value(sheet1, i, file1_col_alm, get_cell_value(sheet2, j, 4))
                        set_cell_value(sheet1, i, file1_col_msg, get_cell_value(sheet2, j, 4))
                        continue
                    if str(get_cell_value(sheet1, i, file1_col_template))[0:4] == 'Dm10':
                        if get_cell_value(sheet1, i, file1_col_event) == '1' \
                                and get_cell_value(sheet1, i, file1_col_prio) == '900':
                            set_cell_value(sheet1, i, file1_col_alm, get_cell_value(sheet2, j, 3))
                            set_cell_value(sheet1, i, file1_col_msg, get_cell_value(sheet2, j, 3))
                            continue
                        set_cell_value(sheet1, i, file1_col_alm, get_cell_value(sheet2, j, 4))
                        set_cell_value(sheet1, i, file1_col_msg, get_cell_value(sheet2, j, 4))
                        continue
                    if str(get_cell_value(sheet1, i, file1_col_template)) != 'Am100':
                        set_cell_value(sheet1, i, file1_col_alm, get_cell_value(sheet2, j, 4))
                        set_cell_value(sheet1, i, file1_col_msg, get_cell_value(sheet2, j, 4))


def am100_clean_devicetags(sheet1, file1_col_template,  file1_col_devicetag2, file1_col_devicetag3, file1_col_devicetag4,
                           file1_col_devicetag5, file1_col_devicetag6, file1_col_devicetag7, file1_col_devicetag8,
                           file1_col_devicetag9, file1_col_devicetag10, max_rows1):
    """
    Function 'am100_clean_devicetag' cleaning devicetag 2 --> 10 for Am100 loops only and prepare to dynamic filling.
    ----------
    """""
    for i in range(2, max_rows1+1):
        if get_cell_value(file1_col_template) == 'Am100':
            set_cell_value(sheet1, i, file1_col_devicetag2, 2)
            set_cell_value(sheet1, i, file1_col_devicetag3, 2)
            set_cell_value(sheet1, i, file1_col_devicetag4, 2)
            set_cell_value(sheet1, i, file1_col_devicetag5, 2)
            set_cell_value(sheet1, i, file1_col_devicetag6, 2)
            set_cell_value(sheet1, i, file1_col_devicetag7, 2)
            set_cell_value(sheet1, i, file1_col_devicetag8, 2)
            set_cell_value(sheet1, i, file1_col_devicetag9, 2)
            set_cell_value(sheet1, i, file1_col_devicetag10, 2)


