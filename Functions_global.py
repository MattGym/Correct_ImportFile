from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from tkinter import messagebox
import openpyxl


def clean_all(sheet, col_pcs, col_signed_unsigned, col_signal, max_col, max_row):
    """
    Cleaning cells from description and color
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    col_pcs : int
        pcs column number (skip)
    col_signed_unsigned : int
        signed_unsigned column number (skip)
    col_signal :
        signal column number (skip)
    max_col : int
        number of columns in active_sheet
    max_row : int
        number of rows in active_sheet
    """
    for i in range(1, max_row+1):
        for j in range(1, max_col+1):
            if j != col_pcs:
                set_cell_color(sheet, i, j, 'n')
                set_cell_comment(sheet, i, j, '', False, True)


def get_col_no(sheet, desc, max_col) -> int:
    """
    According to given active_sheet name and searched column description (str), function
    returns column number where that string is.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    desc : str
        searched string
    max_col : int
        number of not empty columns
    """
    for s in range(1, max_col+1):
        if str(desc) == str(sheet.cell(row=1, column=s).value):
            return s
    msg_text = 'Cannot find ' + str(desc) + ' in connected data base.'
    messagebox.showinfo(title='Message', message=msg_text)
    return 0


def get_cell_color(sheet, row, col):
    """
    Checking is cell color is a function one.
    Output : str - empty, red, gree, blue, grey
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    """
    color = sheet.cell(row=row, column=col).fill.fgColor.rgb
    out = 'empty'
    if color == '00E74C3C':
        out = 'red'
    if color == '002ECC71':
        out = 'green'
    if color == '002E80CC':
        out = 'blue'
    if color == '008A8A8A':
        out = 'grey'
    return out


def set_cell_color(sheet, row, col, color_rgn='n'):
    """
    Change cell background color on 'r'-red, 'g'-green, 'b'-blue, 'grey'-grey, 'n'-none. ('row_name'-row name, 'row_even'-row even, 'row_odd'=row odd)
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    color_rgn : str, optional
        (OPTIONAL) color_rgn = 'r'-red; 'g'-green; 'grey'-grey;'n' or empty - NONE
    """
    if row > 0 and col > 0:
        if color_rgn == 'r':        # red
            fg_color = 'E74C3C'
            pattern = PatternFill(patternType='solid', fgColor=fg_color)
            sheet.cell(row=row, column=col).fill = pattern
        if color_rgn == 'g':        # green
            fg_color = '2ECC71'
            pattern = PatternFill(patternType='solid', fgColor=fg_color)
            sheet.cell(row=row, column=col).fill = pattern
        if color_rgn == 'b':        # blue
            fg_color = '2E80CC'
            pattern = PatternFill(patternType='solid', fgColor=fg_color)
            sheet.cell(row=row, column=col).fill = pattern
        if color_rgn == 'grey':     # grey
            fg_color = '8A8A8A'
            pattern = PatternFill(patternType='solid', fgColor=fg_color)
            sheet.cell(row=row, column=col).fill = pattern
        if color_rgn == 'row_name':  # Light blue (1st row)
            fg_color = '6FA8DC'
            pattern = PatternFill(patternType='solid', fgColor=fg_color)
            sheet.cell(row=row, column=col).fill = pattern
        if color_rgn == 'row_even':  # Darker blue
            fg_color = '9FC5E8'
            pattern = PatternFill(patternType='solid', fgColor=fg_color)
            sheet.cell(row=row, column=col).fill = pattern
        if color_rgn == 'row_odd':  # Lighter blue
            fg_color = 'CFE2F3'
            pattern = PatternFill(patternType='solid', fgColor=fg_color)
            sheet.cell(row=row, column=col).fill = pattern
        if color_rgn == 'n':        # none color
            pattern = PatternFill(patternType=None)
            sheet.cell(row=row, column=col).fill = pattern
    else:
        messagebox.showinfo(title='Message',
                            message='Row or column cannot be 0, please check column names in mapping file')


def get_cell_value(sheet, row, col):
    """
    Returns value from cell in active_sheet at specific row & column position.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    """
    if row > 0 and col > 0:
        return sheet.cell(row=row, column=col).value
    else:
        messagebox.showinfo(title='Message', message='Row or column cannot be 0, please check column names in mapping file')
        return None


def set_cell_value(sheet, row, col, val, typ=0):
    """
    Set value at cell in active_sheet at specific row & column position.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    val : any
        value
    typ : int, optional
        (OPTIONAL) type of data (0 or empty - float/int; 1 - force string; 2 - None)
    """
    if typ == 0:
        sheet.cell(row=row, column=col).value = float(val)
    elif typ == 1:
        sheet.cell(row=row, column=col).value = str(val)
    elif typ == 2:
        sheet.cell(row=row, column=col).value = None
    elif typ == 3:
        sheet.cell(row=row, column=col).value = int(val)


def set_cell_comment(sheet, row, col, commentary, add=False, delete=False):
    """
    Function add 'commentary' to specified cell given as (active_sheet, row & column position)
    If optional parameter add=True then function add commentary to existing one.
    In another way removes old commentary and add a new one.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    commentary : str
        commentary that will be added to specified cell
    add : bool, optional
        (OPTIONAL) add=False or NONE - function swap commentary into new on , add=True - add second commentary
    delete : bool, optional
        (OPTIONAL) delete=False or NONE - function do nothing with existing commentary, delete=True - remove commentary
    """
    if row > 0 and col > 0:
        if str(sheet.cell(row=row, column=col).comment) == 'None' or add is False:
            comment = Comment(commentary, 'CdA analyzer')
            comment.width = 500
            comment.height = 200
            sheet.cell(row=row, column=col).comment = comment
        else:
            tmp_txt1 = str(str(sheet.cell(row=row, column=col).comment).replace('Comment: ', '')).\
                replace('by CdA analyzer', '')
            comment = Comment(tmp_txt1 + ' ::\n' + commentary, 'CdA analyzer')
            comment.width = 500
            comment.height = 200
            sheet.cell(row=row, column=col).comment = comment
        if delete is True:
            sheet.cell(row=row, column=col).comment = None
    else:
        messagebox.showinfo(title='Message',
                            message='Row or column cannot be 0, please check column names in mapping file')


def add_column(sheet, max_col, max_row, description, anchor=0):
    """
    Adding column at the beginning or end of sheet.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    max_col : int
        number of columns
    max_row : int
        number of rows
    description : str
        column description
    anchor : int
        anchor place for new column (0 - front, 1 - end)
    """
    if get_col_no(sheet, description, max_col + 1) > 0:
        old_col_no = get_col_no(sheet, description, max_col + 1)
        if anchor == 0:
            sheet.insert_cols(1)
            set_cell_value(sheet, 1, 1, description, 1)
            set_cell_color(sheet, 1, 1, 'grey')
            for j in range(2, max_row + 1):
                if get_cell_value(sheet, j, old_col_no + 1) is not None:
                    set_cell_value(sheet, j, 1, str(get_cell_value(sheet, j, old_col_no + 1)), 1)
            sheet.delete_cols(old_col_no + 1)
        else:
            sheet.insert_cols(max_col + 1)
            set_cell_value(sheet, 1, max_col + 2, description, 1)
            set_cell_color(sheet, 1, max_col + 2, 'grey')
            for j in range(2, max_row + 1):
                if get_cell_value(sheet, j, old_col_no) is not None:
                    set_cell_value(sheet, j, max_col + 2, str(get_cell_value(sheet, j, old_col_no)), 1)
            sheet.delete_cols(old_col_no)
    else:
        if anchor == 0:
            sheet.insert_cols(1)
            set_cell_value(sheet, 1, 1, description, 1)
            set_cell_color(sheet, 1, 1, 'grey')
        else:
            sheet.insert_cols(max_col + 1)
            set_cell_value(sheet, 1, max_col + 1, description, 1)
            set_cell_color(sheet, 1, max_col + 1, 'grey')


def hide_unused_col(sheet, max_col: int, start_hide_after: int):
    """
    Hiding unused columns (not edited columns)
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    max_col : int
        number of columns
    start_hide_after : int
        Start hiding columns after col no
    """
    start_hiding = start_hide_after + 1
    for i in range(start_hiding, max_col + 1):
        hide = True
        color = get_cell_color(sheet, 1, i)
        if color == 'red' or color == 'green' or color == 'blue' or color == 'grey':
            hide = False
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].hidden = hide

